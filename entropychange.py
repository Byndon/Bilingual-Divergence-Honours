import openpyxl
import os
import sys
import math
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


os.chdir("/home/callan/2020Honours/ABM/modelgit/outputs/outs/")


def loadSpreadsheet(filename):
    '''Load the Speradsheet at filename (relative to the above directory) into the program'''
    workbook = openpyxl.load_workbook(filename, data_only=True)
    return workbook


def findChangeOfEntropy(sheet, steps):
    
    eColumnIndex = sheet.max_column
    entropyChangeColumn = openpyxl.utils.cell.get_column_letter(eColumnIndex)
    cellRangeIndicator = entropyChangeColumn + "1" + ":" + entropyChangeColumn + str((steps+2)*110)
    return(cellRangeIndicator)


def main(filename, steps):
    
    dataDict = {"Language": [], "Mode": [], "Monitoring": [], "Weight": [], "delta H": []}
    workbook = loadSpreadsheet(filename)
    for sheet in workbook:
        entropyChangeColumn = findChangeOfEntropy(sheet, steps)

        for chunk in chunks(sheet[entropyChangeColumn], steps + 2):
            sumlist = []
            for cell in chunk:
                if((type(cell[0].value) is float) or (type(cell[0].value) is int)):
                    sumlist.append(cell[0].value)
            dataDict["delta H"].append(math.fsum(sumlist))
            dataDict["Language"].append(sheet.title)

        for row in sheet.iter_rows(min_row=1, max_col=5, max_row=(steps+2)*110):
            for cell in row:
                if(cell.row % 37 == 1):
                    if(cell.column == 2):
                        dataDict["Mode"].append(cell.value)
                    elif(cell.column == 3):
                        dataDict["Monitoring"].append(cell.value)
                    elif(cell.column == 5):
                        dataDict["Weight"].append(cell.value)

    dataframe = pd.DataFrame(data=dataDict)
    print(dataframe)

    outworkbook = openpyxl.Workbook()
    datasheet = outworkbook.create_sheet("Correlates")
    for row in dataframe_to_rows(dataframe, index=True, header=True):
        datasheet.append(row)

    for cell in datasheet["A"] + datasheet[1]:
        cell.style = "Pandas"
    outfilename = "correlates" + filename
    outworkbook.save(filename=outfilename)


fileList = os.listdir()

for i in fileList:
    main(i, 35)
